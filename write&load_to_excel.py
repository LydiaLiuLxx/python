from openpyxl import Workbook
#--------creat & save ----------
wb = Workbook()
sheet = wb.active
sheet.title='ID'
sheet.cell(row=1,column=1).value='ID'

for row in range(1,40):
    sheet.append(range(row))

#sheet2=wb.create_sheet(title='name')

wb.save('write_to_excel1.xlsx')

#---------load &print-------
from openpyxl import load_workbook
wb = load_workbook(filename = 'write_to_excel1.xlsx')
sheet_ranges = wb['ID']
print(sheet_ranges['D18'].value)
